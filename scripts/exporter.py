"""
Export module for Job Search Tool.

Handles saving job search results to CSV and formatted Excel files.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from logger import get_logger

# Score threshold for highlighting in Excel
HIGH_SCORE_THRESHOLD = 30

# Maximum column width in Excel auto-sizing
MAX_COLUMN_WIDTH = 50


def _sanitize_excel_value(value: object) -> object:
    """Escape values that could trigger Excel formula injection.

    Cell values starting with ``=`` or ``@`` are always prefixed.
    Values starting with ``+`` or ``-`` are only prefixed when followed
    by a character that could form a formula (digit, letter, or another
    operator), which avoids corrupting legitimate text like job
    descriptions starting with a dash.
    """
    if not isinstance(value, str) or not value:
        return value
    first = value[0]
    if first in ("=", "@"):
        return "'" + value
    if (
        first in ("+", "-")
        and len(value) > 1
        and (value[1].isdigit() or value[1] in ("=", "+", "-", "@"))
    ):
        return "'" + value
    return value


def _sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of *df* with text columns sanitized against formula injection."""
    sanitized = df.copy()
    text_cols = sanitized.select_dtypes(include=["object"]).columns
    for col in text_cols:
        sanitized[col] = sanitized[col].map(
            lambda v: _sanitize_excel_value(v) if pd.notna(v) else v
        )
    return sanitized


def dataframe_to_csv_bytes(jobs_df: pd.DataFrame) -> bytes:
    """
    Render CSV bytes with spreadsheet-formula sanitization applied.

    Args:
        jobs_df: DataFrame to serialize.

    Returns:
        UTF-8 encoded CSV bytes.
    """
    safe_df = _sanitize_dataframe_for_excel(jobs_df)
    return safe_df.to_csv(index=False).encode("utf-8")


def _write_excel_workbook(jobs_df: pd.DataFrame, output: str | BytesIO) -> None:
    """Write a formatted Excel workbook to a path or in-memory buffer."""
    safe_df = _sanitize_dataframe_for_excel(jobs_df)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False, sheet_name="Jobs")
        worksheet = writer.sheets["Jobs"]

        # Format header row
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, _column_title in enumerate(jobs_df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col_num, column in enumerate(jobs_df.columns, 1):
            # Handle empty columns or all-NaN columns gracefully
            col_max = jobs_df[column].astype(str).map(len).max()
            if pd.isna(col_max):
                col_max = 0
            max_length = max(int(col_max), len(str(column)))
            adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
            worksheet.column_dimensions[
                get_column_letter(col_num)
            ].width = adjusted_width

        # Make URLs clickable
        if "job_url" in jobs_df.columns:
            url_col = jobs_df.columns.get_loc("job_url") + 1
            for row_num in range(2, len(jobs_df) + 2):
                cell = worksheet.cell(row=row_num, column=url_col)
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Conditional formatting for relevance score
        if "relevance_score" in jobs_df.columns:
            score_col = jobs_df.columns.get_loc("relevance_score") + 1
            high_score_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            for row_num in range(2, len(jobs_df) + 2):
                cell = worksheet.cell(row=row_num, column=score_col)
                try:
                    if (
                        cell.value is not None
                        and float(cell.value) >= HIGH_SCORE_THRESHOLD
                    ):
                        cell.fill = high_score_fill
                except (ValueError, TypeError):
                    # Skip cells with non-numeric values
                    pass


def dataframe_to_excel_bytes(jobs_df: pd.DataFrame) -> bytes:
    """
    Render a formatted Excel workbook in memory.

    Args:
        jobs_df: DataFrame to serialize.

    Returns:
        Excel workbook bytes, or ``b""`` for empty data.
    """
    if len(jobs_df) == 0:
        return b""

    buffer = BytesIO()
    _write_excel_workbook(jobs_df, buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_dataframe(
    jobs_df: pd.DataFrame,
    output_dir: Path,
    basename: str,
    fmt: str,
) -> Path:
    """Export ``jobs_df`` to ``output_dir`` on demand.

    Creates ``output_dir`` lazily — no side effects unless this helper is
    actually called. Always applies spreadsheet-formula sanitization.

    Args:
        jobs_df: DataFrame to serialize.
        output_dir: Target directory (created if missing).
        basename: File name stem (timestamp is appended).
        fmt: Either ``"csv"`` or ``"excel"``.

    Returns:
        Absolute path of the written file.
    """
    if fmt not in ("csv", "excel"):
        raise ValueError(f"fmt must be 'csv' or 'excel', got {fmt!r}")

    logger = get_logger("exporter")
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if fmt == "csv":
        out = output_dir / f"{basename}_{timestamp}.csv"
        safe_df = _sanitize_dataframe_for_excel(jobs_df)
        safe_df.to_csv(out, index=False)
        logger.info("Exported CSV: %s", out)
        return out

    out = output_dir / f"{basename}_{timestamp}.xlsx"
    _write_excel_workbook(jobs_df, str(out))
    logger.info("Exported Excel: %s", out)
    return out
