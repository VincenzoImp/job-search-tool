#!/usr/bin/env python3
"""
Job Search Tool - Automated Job Aggregation.

Aggregates job listings from multiple job boards using JobSpy.
Features parallel execution, throttling, retry logic, SQLite persistence, and configurable scoring.
"""

from __future__ import annotations

import hashlib
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any

import pandas as pd
from jobspy import scrape_jobs
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

from config import Config, get_config
from database import get_database
from logger import ProgressLogger, get_logger, log_section, setup_logging
from models import Job, SearchSummary


# Thread-safe lock for shared data structures
_jobs_lock = threading.Lock()


class ThrottledExecutor:
    """ThreadPoolExecutor wrapper with per-site rate limiting."""

    def __init__(self, config: Config):
        """
        Initialize throttled executor.

        Args:
            config: Configuration object with throttling settings.
        """
        self.config = config
        self.site_locks: dict[str, threading.Lock] = {}
        self.site_last_request: dict[str, float] = {}
        self._lock = threading.Lock()
        self._logger = get_logger("throttle")

    def _get_site_lock(self, site: str) -> threading.Lock:
        """Get or create a lock for a specific site."""
        with self._lock:
            if site not in self.site_locks:
                self.site_locks[site] = threading.Lock()
                self.site_last_request[site] = 0
            return self.site_locks[site]

    def throttled_search(
        self,
        query: str,
        location: str,
        config: Config,
    ) -> tuple[str, str, pd.DataFrame | None, str | None]:
        """
        Execute search with throttling based on configured sites.

        Since JobSpy internally searches all configured sites in parallel,
        we throttle based on the slowest (most restrictive) site delay
        to ensure we don't overwhelm any individual site.

        Args:
            query: Search term.
            location: Location to search.
            config: Configuration object.

        Returns:
            Tuple of (query, location, results_df, error_message).
        """
        if not config.throttling.enabled:
            # No throttling, execute immediately
            return search_single_query(query, location, config)

        # Get the maximum delay across all configured sites
        # This ensures we don't overwhelm the most restrictive site
        max_delay = config.throttling.default_delay
        for site in config.search.sites:
            site_delay = config.throttling.site_delays.get(
                site.lower(), config.throttling.default_delay
            )
            max_delay = max(max_delay, site_delay)

        # Use a global lock for throttling since we're searching all sites at once
        global_lock = self._get_site_lock("_global")

        with global_lock:
            # Calculate required delay
            now = time.time()
            elapsed = now - self.site_last_request.get("_global", 0)

            # Apply jitter to the delay
            if config.throttling.jitter > 0:
                import random

                jitter_amount = max_delay * config.throttling.jitter
                actual_delay = max_delay + random.uniform(-jitter_amount, jitter_amount)
            else:
                actual_delay = max_delay

            if elapsed < actual_delay:
                sleep_time = actual_delay - elapsed
                self._logger.debug(
                    f"Throttling: waiting {sleep_time:.2f}s before next request"
                )
                time.sleep(sleep_time)

            # Update last request time
            self.site_last_request["_global"] = time.time()

        # Execute search outside the lock
        return search_single_query(query, location, config)


def _normalize_text(text: str) -> str:
    """
    Normalize text for fuzzy matching.

    Handles common character variations (ü->u, ö->o, etc.) and lowercases.
    """
    if not text:
        return ""

    # Lowercase
    text = text.lower()

    # Common character normalizations
    replacements = {
        "ü": "u",
        "ö": "o",
        "ä": "a",
        "ß": "ss",
        "é": "e",
        "è": "e",
        "ê": "e",
        "à": "a",
        "â": "a",
        "î": "i",
        "ô": "o",
        "û": "u",
        "ç": "c",
        "ñ": "n",
    }
    for orig, replacement in replacements.items():
        text = text.replace(orig, replacement)

    return text


def _extract_words(text: str) -> list[str]:
    """Extract words from text, filtering out common stop words."""
    if not text:
        return []

    # Normalize and split
    normalized = _normalize_text(text)
    words = re.findall(r"\b[a-z0-9+#]+\b", normalized)

    # Filter stop words (common words that don't add value to matching)
    stop_words = {
        "a", "an", "the", "and", "or", "but", "in", "on", "at", "to", "for",
        "of", "with", "by", "from", "as", "is", "was", "are", "were", "been",
        "be", "have", "has", "had", "do", "does", "did", "will", "would",
        "could", "should", "may", "might", "must", "shall", "can", "need",
        "that", "this", "these", "those", "it", "its", "we", "you", "they",
        "i", "he", "she", "who", "what", "which", "where", "when", "why", "how",
    }

    return [w for w in words if w not in stop_words and len(w) > 1]


def _fuzzy_word_match(word: str, text: str, min_similarity: int) -> bool:
    """
    Check if a word fuzzy-matches anywhere in the text.

    Uses token set ratio for flexibility with word order and partial matches.
    """
    normalized_word = _normalize_text(word)
    normalized_text = _normalize_text(text)

    # Direct substring match (fastest check)
    if normalized_word in normalized_text:
        return True

    # Extract words from text and check each one
    text_words = _extract_words(text)
    for text_word in text_words:
        # Use partial ratio for handling substrings and fuzz ratio for typos
        similarity = max(
            fuzz.ratio(normalized_word, text_word),
            fuzz.partial_ratio(normalized_word, text_word),
        )
        if similarity >= min_similarity:
            return True

    return False


def _get_job_text(row: pd.Series) -> str:
    """Concatenate all relevant job fields for matching."""
    fields = ["title", "description", "company", "location"]
    parts = []
    for field in fields:
        value = row.get(field)
        if value and pd.notna(value):
            parts.append(str(value))
    return " ".join(parts)


def fuzzy_post_filter(
    jobs_df: pd.DataFrame,
    query: str,
    location: str,
    config: Config,
) -> pd.DataFrame:
    """
    Filter jobs to ensure query terms are present in job data.

    Uses fuzzy matching to handle typos and character variations.

    Args:
        jobs_df: DataFrame with job results.
        query: Original search query.
        location: Search location.
        config: Configuration with post_filter settings.

    Returns:
        Filtered DataFrame with only matching jobs.
    """
    if not config.post_filter.enabled:
        return jobs_df

    if jobs_df is None or len(jobs_df) == 0:
        return jobs_df

    logger = get_logger("post_filter")
    min_similarity = config.post_filter.min_similarity

    # Extract query terms to check
    query_terms = _extract_words(query)

    # Also check location if enabled
    if config.post_filter.check_location and location.lower() != "remote":
        location_terms = _extract_words(location)
    else:
        location_terms = []

    if not query_terms and not location_terms:
        return jobs_df

    logger.debug(
        f"Post-filtering: query_terms={query_terms}, location_terms={location_terms}"
    )

    # Filter rows
    matching_indices = []
    for idx, row in jobs_df.iterrows():
        job_text = _get_job_text(row)

        # Check if all query terms match
        if config.post_filter.check_query_terms:
            query_match = all(
                _fuzzy_word_match(term, job_text, min_similarity)
                for term in query_terms
            )
        else:
            query_match = True

        # Check if location matches (at least one location term should match)
        if location_terms:
            location_match = any(
                _fuzzy_word_match(term, job_text, min_similarity)
                for term in location_terms
            )
        else:
            location_match = True

        if query_match and location_match:
            matching_indices.append(idx)

    filtered_df = jobs_df.loc[matching_indices].copy()

    filtered_count = len(jobs_df) - len(filtered_df)
    if filtered_count > 0:
        logger.debug(
            f"Post-filter removed {filtered_count} jobs "
            f"(kept {len(filtered_df)}/{len(jobs_df)})"
        )

    return filtered_df


def calculate_relevance_score(row: pd.Series, config: Config) -> int:
    """
    Calculate relevance score based entirely on user configuration.

    The scoring system is fully dynamic: it iterates over all keyword categories
    defined in config.scoring.keywords and applies the corresponding weight from
    config.scoring.weights. No hardcoded categories or keywords.

    Args:
        row: DataFrame row with job data.
        config: Configuration with scoring weights and keywords.

    Returns:
        Relevance score as integer (sum of matched category weights).

    Example config:
        scoring:
          weights:
            primary_skills: 20
            technologies: 15
          keywords:
            primary_skills:
              - "software engineer"
              - "backend"
            technologies:
              - "python"
              - "docker"
    """
    # Build searchable text from job fields
    text = " ".join(
        str(row.get(field, "") or "")
        for field in ("title", "description", "company", "location")
    ).lower()

    if not text.strip():
        return 0

    score = 0
    weights = config.scoring.weights
    keywords = config.scoring.keywords

    # Iterate over all keyword categories defined in configuration
    for category, keyword_list in keywords.items():
        if not keyword_list:
            continue

        # Check if any keyword from this category matches
        if any(keyword.lower() in text for keyword in keyword_list):
            # Get weight for this category (default 0 if not specified)
            weight = weights.get(category, 0)
            score += weight

    return score


def search_single_query(
    query: str,
    location: str,
    config: Config,
) -> tuple[str, str, pd.DataFrame | None, str | None]:
    """
    Execute a single search query with retry logic.

    Args:
        query: Search term.
        location: Location to search.
        config: Configuration object.

    Returns:
        Tuple of (query, location, results_df, error_message).
    """
    logger = get_logger("search")

    @retry(
        retry=retry_if_exception_type((ConnectionError, TimeoutError)),
        stop=stop_after_attempt(config.retry.max_attempts),
        wait=wait_exponential(
            multiplier=config.retry.base_delay,
            exp_base=config.retry.backoff_factor,
        ),
        reraise=True,
    )
    def _do_search() -> pd.DataFrame:
        # NOTE: For Indeed, using hours_old disables job_type filtering (JobSpy limitation)
        # We prioritize date filtering over job type for broader results
        return scrape_jobs(
            # Core parameters
            site_name=config.search.sites,
            search_term=query,
            location=location,
            results_wanted=config.search.results_wanted,
            hours_old=config.search.hours_old,
            country_indeed=config.search.country_indeed,
            # JobSpy core parameters
            distance=config.search.distance,
            is_remote=config.search.is_remote,
            easy_apply=config.search.easy_apply,
            offset=config.search.offset,
            # Output format parameters
            enforce_annual_salary=config.search.enforce_annual_salary,
            description_format=config.search.description_format,
            verbose=config.search.verbose,
            # LinkedIn-specific parameters
            linkedin_fetch_description=config.search.linkedin_fetch_description,
            linkedin_company_ids=config.search.linkedin_company_ids,
            # Google Jobs-specific parameters
            google_search_term=config.search.google_search_term,
            # Network/Proxy parameters
            proxies=config.search.proxies,
            ca_cert=config.search.ca_cert,
            user_agent=config.search.user_agent,
        )

    try:
        jobs = _do_search()

        if jobs is not None and len(jobs) > 0:
            jobs["search_query"] = query
            jobs["search_location"] = location
            jobs["search_date"] = datetime.now().strftime("%Y-%m-%d")

            # Apply fuzzy post-filter to validate results match query terms
            original_count = len(jobs)
            jobs = fuzzy_post_filter(jobs, query, location, config)

            if len(jobs) < original_count:
                logger.debug(
                    f"Post-filter: {query} @ {location}: "
                    f"{original_count} -> {len(jobs)} jobs"
                )

            if len(jobs) == 0:
                return query, location, None, None

            return query, location, jobs, None
        else:
            return query, location, None, None

    except Exception as e:
        error_msg = str(e)
        logger.warning(f"Query failed: {query} @ {location}: {error_msg}")
        return query, location, None, error_msg


def search_jobs(config: Config) -> tuple[pd.DataFrame | None, SearchSummary]:
    """
    Search for jobs using parallel execution with throttling.

    Args:
        config: Configuration object.

    Returns:
        Tuple of (combined_jobs_df, search_summary).
    """
    logger = get_logger("search")
    summary = SearchSummary()

    log_section(logger, "SEARCHING FOR JOBS")

    queries = config.get_all_queries()
    locations = config.search.locations

    # Build list of all search tasks
    tasks = [(q, loc) for loc in locations for q in queries]
    summary.total_queries = len(tasks)

    logger.info(f"Total search tasks: {len(tasks)}")
    logger.info(f"Queries: {len(queries)}, Locations: {len(locations)}")
    logger.info(f"Parallel workers: {config.parallel.max_workers}")

    # Log throttling status
    if config.throttling.enabled:
        # Calculate effective delay (max of configured sites)
        max_delay = config.throttling.default_delay
        for site in config.search.sites:
            site_delay = config.throttling.site_delays.get(
                site.lower(), config.throttling.default_delay
            )
            max_delay = max(max_delay, site_delay)
        logger.info(
            f"Throttling enabled: {max_delay:.1f}s delay "
            f"(jitter={config.throttling.jitter:.0%})"
        )
        # Estimate total time
        estimated_time = len(tasks) * max_delay / config.parallel.max_workers
        logger.info(f"Estimated minimum time: {estimated_time / 60:.1f} minutes")
    else:
        logger.info("Throttling disabled")

    all_jobs: list[pd.DataFrame] = []
    seen_jobs: set[str] = set()  # For deduplication

    progress = ProgressLogger(logger, len(tasks), "Job search")

    # Create throttled executor
    throttled_executor = ThrottledExecutor(config)

    # Execute searches in parallel with throttling
    with ThreadPoolExecutor(max_workers=config.parallel.max_workers) as executor:
        futures = {
            executor.submit(throttled_executor.throttled_search, q, loc, config): (
                q,
                loc,
            )
            for q, loc in tasks
        }

        for future in as_completed(futures):
            query, location = futures[future]

            try:
                q, loc, jobs_df, error = future.result()

                if error:
                    summary.failed_queries += 1
                    progress.update(
                        success=False, message=f"FAILED: {query} @ {location}"
                    )
                elif jobs_df is not None and len(jobs_df) > 0:
                    summary.successful_queries += 1
                    summary.total_jobs_found += len(jobs_df)

                    # Deduplicate incrementally (all operations under single lock)
                    with _jobs_lock:
                        new_jobs = []
                        for _, row in jobs_df.iterrows():
                            job_key = _generate_job_key(row)
                            if job_key not in seen_jobs:
                                seen_jobs.add(job_key)
                                new_jobs.append(row)

                        if new_jobs:
                            new_df = pd.DataFrame(new_jobs)
                            all_jobs.append(new_df)

                    progress.update(
                        success=True,
                        message=f"Found {len(jobs_df)} jobs: {query} @ {location}",
                    )
                else:
                    summary.successful_queries += 1
                    progress.update(
                        success=True, message=f"No results: {query} @ {location}"
                    )

            except Exception as e:
                summary.failed_queries += 1
                logger.error(f"Unexpected error for {query} @ {location}: {e}")
                progress.update(success=False, message=f"ERROR: {query}")

    progress.summary()

    if not all_jobs:
        logger.warning("No jobs found in any search.")
        summary.finish()
        return None, summary

    # Combine all results
    combined_jobs = pd.concat(all_jobs, ignore_index=True)
    summary.unique_jobs = len(combined_jobs)

    logger.info(f"Total unique jobs after deduplication: {len(combined_jobs)}")

    summary.finish()
    return combined_jobs, summary


def _generate_job_key(row: pd.Series) -> str:
    """Generate unique key for job deduplication."""
    identifier = f"{row.get('title', '')}|{row.get('company', '')}|{row.get('location', '')}".lower()
    return hashlib.sha256(identifier.encode()).hexdigest()[:16]


def filter_relevant_jobs(jobs_df: pd.DataFrame, config: Config) -> pd.DataFrame:
    """
    Filter jobs based on relevance score.

    Args:
        jobs_df: DataFrame with all jobs.
        config: Configuration with scoring settings.

    Returns:
        Filtered DataFrame with relevant jobs.
    """
    logger = get_logger("filter")

    log_section(logger, "FILTERING RELEVANT JOBS")

    # Calculate relevance scores
    jobs_df = jobs_df.copy()
    jobs_df["relevance_score"] = jobs_df.apply(
        lambda row: calculate_relevance_score(row, config), axis=1
    )

    # Filter by threshold
    threshold = config.scoring.threshold
    relevant_jobs = jobs_df[jobs_df["relevance_score"] > threshold].copy()
    relevant_jobs = relevant_jobs.sort_values("relevance_score", ascending=False)

    logger.info(f"Score threshold: {threshold}")
    logger.info(f"Relevant jobs found: {len(relevant_jobs)}")

    if len(relevant_jobs) > 0:
        logger.info(f"Highest score: {relevant_jobs['relevance_score'].max()}")
        logger.info(f"Average score: {relevant_jobs['relevance_score'].mean():.1f}")

    return relevant_jobs


def save_results(
    jobs_df: pd.DataFrame,
    config: Config,
    filename_prefix: str = "jobs",
) -> tuple[str, str]:
    """
    Save results to CSV and Excel with formatting.

    Args:
        jobs_df: DataFrame to save.
        config: Configuration object.
        filename_prefix: Prefix for output files.

    Returns:
        Tuple of (csv_path, excel_path).
    """
    logger = get_logger("output")

    # Check if any output is enabled
    if not config.output.save_csv and not config.output.save_excel:
        logger.info("File output disabled (save_csv and save_excel are both false)")
        return "", ""

    results_dir = config.results_path
    results_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = ""
    excel_path = ""

    # Save to CSV
    if config.output.save_csv:
        csv_file = results_dir / f"{filename_prefix}_{timestamp}.csv"
        jobs_df.to_csv(csv_file, index=False)
        logger.info(f"Saved CSV: {csv_file}")
        csv_path = str(csv_file)

    # Save to Excel with formatting
    if config.output.save_excel:
        excel_file = results_dir / f"{filename_prefix}_{timestamp}.xlsx"

        try:
            with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
                jobs_df.to_excel(writer, index=False, sheet_name="Jobs")
                worksheet = writer.sheets["Jobs"]

                # Format header row
                header_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")

                for col_num, column_title in enumerate(jobs_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # Auto-adjust column widths
                for col_num, column in enumerate(jobs_df.columns, 1):
                    max_length = max(
                        jobs_df[column].astype(str).map(len).max(),
                        len(str(column)),
                    )
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_num)].width = (
                        adjusted_width
                    )

                # Make URLs clickable
                if "job_url" in jobs_df.columns:
                    url_col = jobs_df.columns.get_loc("job_url") + 1
                    for row_num in range(2, len(jobs_df) + 2):
                        cell = worksheet.cell(row=row_num, column=url_col)
                        if cell.value and str(cell.value).startswith("http"):
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0563C1", underline="single")

                # Freeze header row
                worksheet.freeze_panes = "A2"

                # Conditional formatting for relevance score
                if "relevance_score" in jobs_df.columns:
                    score_col = jobs_df.columns.get_loc("relevance_score") + 1
                    high_score_fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    for row_num in range(2, len(jobs_df) + 2):
                        cell = worksheet.cell(row=row_num, column=score_col)
                        try:
                            if cell.value is not None and float(cell.value) >= 30:
                                cell.fill = high_score_fill
                        except (ValueError, TypeError):
                            # Skip cells with non-numeric values
                            pass

            logger.info(f"Saved Excel: {excel_file}")
            excel_path = str(excel_file)

        except Exception as e:
            logger.warning(f"Could not save Excel file: {e}")

    return csv_path, excel_path


def print_banner(config: Config) -> None:
    """Print application banner with profile info using logger."""
    logger = get_logger("banner")
    profile = config.profile

    # Truncate long strings to fit banner width
    def truncate(text: str, max_len: int) -> str:
        return text[:max_len] if len(text) <= max_len else text[: max_len - 3] + "..."

    name = truncate(profile.name, 33)
    position = truncate(profile.current_position, 65)
    skills = truncate(profile.skills, 56)
    target = truncate(profile.target, 56)

    banner_lines = [
        "",
        "╔══════════════════════════════════════════════════════════════════════╗",
        f"║             Job Search Tool - {name:<33} ║",
        "║                                                                      ║",
        "║  Profile:                                                            ║",
        f"║  • {position:<65} ║",
        f"║  • Skills: {skills:<56} ║",
        f"║  • Target: {target:<56} ║",
        "╚══════════════════════════════════════════════════════════════════════╝",
        "",
    ]

    for line in banner_lines:
        logger.info(line)


def print_top_jobs(jobs_df: pd.DataFrame, count: int = 10) -> None:
    """Print top N jobs by relevance score."""
    logger = get_logger("results")

    log_section(logger, f"TOP {count} MOST RELEVANT JOBS")

    for idx, (_, row) in enumerate(jobs_df.head(count).iterrows(), 1):
        logger.info(f"\n{idx}. {row['title']}")
        logger.info(f"   Company: {row['company']}")
        logger.info(f"   Location: {row['location']}")
        logger.info(f"   Relevance Score: {row['relevance_score']}")
        if pd.notna(row.get("job_url")):
            logger.info(f"   URL: {row['job_url']}")


def main() -> None:
    """Main execution function."""
    # Load configuration
    config = get_config()

    # Setup logging
    logger = setup_logging(config)

    # Print banner
    print_banner(config)

    # Initialize database
    db = get_database(config)
    db_stats = db.get_statistics()
    logger.info(f"Database: {db_stats['total_jobs']} jobs tracked")

    # Search for jobs
    all_jobs, summary = search_jobs(config)

    if all_jobs is None or len(all_jobs) == 0:
        logger.error("No jobs found. Exiting.")
        return

    # Save all results
    log_section(logger, "SAVING ALL RESULTS")
    save_results(all_jobs, config, "all_jobs")

    # Filter relevant jobs
    relevant_jobs = filter_relevant_jobs(all_jobs, config)
    summary.relevant_jobs = len(relevant_jobs)

    if len(relevant_jobs) == 0:
        logger.warning("No highly relevant jobs found after filtering.")
        logger.info("Tip: Check the 'all_jobs' file for broader results.")
        return

    # Save filtered results
    log_section(logger, "SAVING FILTERED RESULTS")
    save_results(relevant_jobs, config, "relevant_jobs")

    # Save to database and identify new jobs
    new_count, updated_count = db.save_jobs_from_dataframe(relevant_jobs)
    summary.new_jobs = new_count

    logger.info(f"Database updated: {new_count} new, {updated_count} existing")

    # Print top matches
    print_top_jobs(relevant_jobs)

    # Print summary
    log_section(logger, "SEARCH COMPLETE")
    logger.info(f"Duration: {summary.duration_formatted}")
    logger.info(f"Total unique jobs: {summary.unique_jobs}")
    logger.info(f"Highly relevant jobs: {summary.relevant_jobs}")
    logger.info(f"New jobs (first time seen): {summary.new_jobs}")
    logger.info(f"Check the 'results' folder for detailed CSV and Excel files.")


if __name__ == "__main__":
    main()
