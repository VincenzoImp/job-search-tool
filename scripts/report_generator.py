"""
Report generator for Job Search Tool.

Generates formatted reports for notifications and exports.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from models import JobDBRecord


@dataclass
class SearchReport:
    """Container for search report data."""

    timestamp: datetime
    total_jobs: int
    new_jobs: int
    updated_jobs: int
    avg_score: float
    top_jobs: list[JobDBRecord]
    all_new_jobs: list[JobDBRecord]


def generate_text_summary(report: SearchReport) -> str:
    """
    Generate plain text summary of search results.

    Args:
        report: Search report data.

    Returns:
        Plain text summary string.
    """
    lines = [
        "=" * 60,
        "JOB SEARCH TOOL - SEARCH REPORT",
        "=" * 60,
        "",
        f"Timestamp: {report.timestamp.strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "SUMMARY",
        "-" * 40,
        f"Total jobs found:    {report.total_jobs}",
        f"New jobs:            {report.new_jobs}",
        f"Updated jobs:        {report.updated_jobs}",
        f"Average score:       {report.avg_score:.1f}",
        "",
    ]

    if report.new_jobs > 0 and report.top_jobs:
        lines.extend([
            "TOP NEW JOBS",
            "-" * 40,
        ])

        for idx, job in enumerate(report.top_jobs[:10], 1):
            lines.extend([
                f"\n{idx}. {job.title}",
                f"   Company:  {job.company}",
                f"   Location: {job.location}",
                f"   Score:    {job.relevance_score}",
            ])
            if job.job_url:
                lines.append(f"   URL:      {job.job_url}")

    lines.extend([
        "",
        "=" * 60,
    ])

    return "\n".join(lines)


def generate_markdown_summary(report: SearchReport, max_jobs: int = 10) -> str:
    """
    Generate Markdown summary of search results.

    Args:
        report: Search report data.
        max_jobs: Maximum number of jobs to include.

    Returns:
        Markdown formatted summary.
    """
    lines = [
        "# Job Search Tool - Search Report",
        "",
        f"**Timestamp:** {report.timestamp.strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## Summary",
        "",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Total jobs found | {report.total_jobs} |",
        f"| New jobs | {report.new_jobs} |",
        f"| Updated jobs | {report.updated_jobs} |",
        f"| Average score | {report.avg_score:.1f} |",
        "",
    ]

    if report.new_jobs > 0 and report.top_jobs:
        lines.extend([
            "## Top New Jobs",
            "",
        ])

        for idx, job in enumerate(report.top_jobs[:max_jobs], 1):
            lines.append(f"### {idx}. {job.title}")
            lines.append("")
            lines.append(f"- **Company:** {job.company}")
            lines.append(f"- **Location:** {job.location}")
            lines.append(f"- **Score:** {job.relevance_score}")

            if job.is_remote:
                lines.append("- **Remote:** Yes")

            if job.job_url:
                lines.append(f"- [View Job]({job.job_url})")

            lines.append("")

        if len(report.all_new_jobs) > max_jobs:
            remaining = len(report.all_new_jobs) - max_jobs
            lines.append(f"*... and {remaining} more jobs*")
            lines.append("")

    return "\n".join(lines)


def jobs_to_dataframe(jobs: list[JobDBRecord]) -> pd.DataFrame:
    """
    Convert list of JobDBRecord to DataFrame.

    Args:
        jobs: List of job records.

    Returns:
        DataFrame with job data.
    """
    if not jobs:
        return pd.DataFrame()

    data = []
    for job in jobs:
        data.append({
            "title": job.title,
            "company": job.company,
            "location": job.location,
            "relevance_score": job.relevance_score,
            "site": job.site,
            "job_type": job.job_type,
            "is_remote": job.is_remote,
            "job_level": job.job_level,
            "date_posted": job.date_posted,
            "min_amount": job.min_amount,
            "max_amount": job.max_amount,
            "currency": job.currency,
            "job_url": job.job_url,
            "first_seen": job.first_seen,
            "last_seen": job.last_seen,
        })

    return pd.DataFrame(data)


def generate_excel_report(jobs: list[JobDBRecord]) -> BytesIO:
    """
    Generate Excel report from job records.

    Args:
        jobs: List of job records.

    Returns:
        BytesIO buffer containing Excel file.
    """
    df = jobs_to_dataframe(jobs)

    if df.empty:
        # Return empty Excel with headers only
        df = pd.DataFrame(columns=[
            "title", "company", "location", "relevance_score",
            "site", "job_type", "is_remote", "job_level",
            "date_posted", "job_url"
        ])

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="New Jobs")
        worksheet = writer.sheets["New Jobs"]

        # Format header row
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            if len(df) > 0:
                max_length = max(
                    df[column].astype(str).map(len).max(),
                    len(str(column)),
                )
            else:
                max_length = len(str(column))
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Make URLs clickable
        if "job_url" in df.columns and len(df) > 0:
            url_col = df.columns.get_loc("job_url") + 1
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=url_col)
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")

        # Highlight high scores
        if "relevance_score" in df.columns and len(df) > 0:
            score_col = df.columns.get_loc("relevance_score") + 1
            high_score_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=score_col)
                try:
                    if cell.value is not None and float(cell.value) >= 30:
                        cell.fill = high_score_fill
                except (ValueError, TypeError):
                    pass

        # Freeze header row
        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    return buffer
